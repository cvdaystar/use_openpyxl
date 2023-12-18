import os, sys, openpyxl
import json, re, argparse

from openpyxl.utils import get_column_letter, column_index_from_string

class Transformer:

    def __init__(self, config_file):
        
        with open(config_file, 'r') as f:
            self.config = json.loads(f.read())

    def process(self, ws, sheet_config):

        print(ws.title)
        print(sheet_config)

        header_row = sheet_config['header_row']
        data_init_row = sheet_config['data_init_row']

        # show all headers
        headers = []
        for column in range(1, ws.max_column + 1):
            header_value = ws.cell(row=header_row, column=column).value
            headers.append(str(header_value))
            # print('header: {hv} on column {cl}'.format(hv=header_value, cl=get_column_letter(column)))        
        print('|'.join(headers))

        # show all values (by row)
        for row in range(data_init_row, ws.max_row + 1):
            row_data = []
            for column in range(1, ws.max_column + 1):
                data = ws.cell(row=row, column=column).value
                row_data.append(str(data))
                # print('data: {d} on row: {rl}, column {cl}'.format(d=str(data), rl=row, cl=get_column_letter(column)))
            print('|'.join(row_data))

    def run(self, excel_file):
        wb = openpyxl.load_workbook(filename = excel_file, data_only=True)
        sheet_names = wb.get_sheet_names()
        for sheet_name in sheet_names:
            for ws_name_pattern, sheet_config in self.config.items():
                if re.match(ws_name_pattern, sheet_name) != None:
                    ws = wb.get_sheet_by_name(sheet_name)
                    self.process(ws=ws, sheet_config=sheet_config)
    

if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', help='target filename', default=None) 
    args = parser.parse_args()

    if args.file == None:
        raise Exception('please provide target file, ex: python main.py -f xxx.xlsx')

    Transformer(config_file='setting.json').run(args.file)